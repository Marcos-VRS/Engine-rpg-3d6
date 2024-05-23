class BancoDeDadosPersonagem:

    def criar_novo_personagem(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook, Workbook

        # VARIÁVEIS
        nome_do_arquivo = "Fichas dos personagens.xlsx"
        nome_da_pagina = "Temporário"

        # VERIFICANDO SE O ARQUIVO JÁ EXISTE
        if os.path.exists(nome_do_arquivo):
            banco_de_dados_fichas = load_workbook(nome_do_arquivo)
        else:  # CRIA E SALVA INICIALMENTE SE NÃO EXISTIR
            banco_de_dados_fichas = Workbook()
            banco_de_dados_fichas.save(nome_do_arquivo)

        # CRIA UMA NOVA PAGINA
        banco_de_dados_fichas.create_sheet(nome_da_pagina)

        # SELECIONA A NOVA PÁGINA
        pagina = banco_de_dados_fichas[nome_da_pagina]

        # CRIA UMA FICHA TEMPORÁRIA COM O TEMPLATE ABAIXO
        template_ficha = [
            ["Nome: ", "", "Player: ", ""],
            ["Total de pontos: ", ""],
            ["Total de desvantagens: ", ""],
            ["XP guardado: ", ""],
            ["", "", ""],
            ["", "", ""],
            ["ST: ", "valorst", "xpst"],
            ["DX: ", "valordx", "xpdx"],
            ["IQ: ", "valoriq", "xpiq"],
            ["HT: ", "valorht", "xpht"],
            ["", "", ""],
            ["", "", ""],
            ["HP: ", "valorhp"],
            ["FP: ", "valorfp"],
            ["", "", ""],
            ["Swing: ", "valorsw", "d6"],
            ["Thrusting: ", "valorthr", "d6"],
            ["Basic Lift(lbs/kg)", "valor", "lbs", "/", "valor", "kg"],
            ["Encumbrance: ", "valorencumbrance"],
            ["RD: ", "vamorrd"],
            ["", "", ""],
            ["Basic Speed: ", "vamorbs"],
            ["Basic Move: ", "valorbm", "m/s"],
            ["", "", ""],
            ["Will: ", "valorwill"],
            ["Per:", "valorpercepção"],
            ["Stress: ", "-", "valorstress"],
            ["", "", ""],
            ["Dodge: ", "valordodge"],
            ["Parry weapon: ", "valorparryw"],
            ["Parry unarmed: ", "valorparryu"],
            ["", "", ""],
            ["", "", ""],
            ["ADVANTAGE ", "xpvantagens"],
            ["nomevantagem", "xpvantagem"],
            ["", "", ""],
            ["", "", ""],
            ["DISADVANTAGE ", ""],
            ["", "", ""],
            ["", "", ""],
            ["SKILLS ", "xpskill", "", "C", "/", "NH"],
            ["nome", "custo", "/", "nh"],
            ["", "", ""],
            ["", "", ""],
            ["BACKGROUND: "],
            ["história do personagemaqui"],
        ]

        # INSERE O TEMPLATE NA PÁGINA
        for linha_index, linha_data in enumerate(template_ficha, start=1):
            for coluna_index, valor_celula in enumerate(linha_data, start=1):
                pagina.cell(row=linha_index, column=coluna_index, value=valor_celula)

        # SALVA O ARQUIVO
        banco_de_dados_fichas.save(nome_do_arquivo)

    def carregar_personagem(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def salvar_personagem(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def editar_personagem(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def deletar_personagem(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def mostrar_ficha(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def mostrar_atributos(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def mostrar_subatributos(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def mostrar_pericias(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def mostrar_vantagens(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def mostrar_desvantagens(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def mostrar_dados_perfil(self):

        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook

        # VARIÁVEIS
        banco_de_dados_fichas = load_workbook("Fichas dos personagens.xlsx")

    def deletar_personagem_temporario(self): ...


class BancoDeDadosCampanha:
    def criar_nova_campanha(self, nome, tl, pontos, desvantagens):
        # IMPORTS
        import os
        import openpyxl
        from openpyxl import load_workbook, Workbook

        # VARIÁVEIS
        nome_do_arquivo = "Campanhas.xlsx"
        nome_da_pagina = f"{nome} - {pontos}"

        # VERIFICANDO SE O ARQUIVO JÁ EXISTE
        if os.path.exists(nome_do_arquivo):
            banco_de_dados_fichas = load_workbook(nome_do_arquivo)
        else:  # CRIA E SALVA INICIALMENTE SE NÃO EXISTIR
            banco_de_dados_fichas = Workbook()
            banco_de_dados_fichas.save(nome_do_arquivo)

        # CRIA UMA NOVA PAGINA
        banco_de_dados_fichas.create_sheet(nome_da_pagina)

        # SELECIONA A NOVA PÁGINA
        pagina = banco_de_dados_fichas[nome_da_pagina]

        # ESTRUTURA DOS DADOS
        campanha = [
            ["NOME: ", nome],
            ["TL: ", tl],
            ["PONTOS: ", pontos],
            ["DESVANTAGENS: ", desvantagens],
        ]

        # INSERE O TEMPLATE NA PÁGINA
        for linha_index, linha_data in enumerate(campanha, start=1):
            for coluna_index, valor_celula in enumerate(linha_data, start=1):
                pagina.cell(row=linha_index, column=coluna_index, value=valor_celula)

        # SALVA O ARQUIVO
        banco_de_dados_fichas.save(nome_do_arquivo)

    def carregar_campanha(self): ...
    def deletar_camapanha(self): ...
